"""
MODULE 1 — Data Ingestion & Validation
Loads job data from CSV/XLSX, validates it, and outputs clean records.
"""
import csv
import logging
import os
from datetime import date, datetime
from typing import Optional

try:
    import openpyxl
    XLSX_SUPPORTED = True
except ImportError:
    XLSX_SUPPORTED = False

REQUIRED_FIELDS = [
    "job_id", "utility_owner", "contractor", "scope_type", "region",
    "start_date", "planned_end_date", "status",
    "markout_required", "markout_issues", "inspections_failed", "crew_type",
]

VALID_STATUSES = {"Open", "In Progress", "Completed"}


def setup_logging(log_path: str = "logs/validation_errors.log"):
    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    logging.basicConfig(
        filename=log_path,
        level=logging.WARNING,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    return logging.getLogger("prg_ops.validation")


def parse_date(val: str) -> Optional[date]:
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(val.strip(), fmt).date()
        except (ValueError, AttributeError):
            pass
    return None


def parse_bool(val) -> Optional[bool]:
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        if val.strip().lower() in ("true", "yes", "1"):
            return True
        if val.strip().lower() in ("false", "no", "0"):
            return False
    return None


def load_jobs(file_path: str) -> list[dict]:
    """Load job records from CSV or XLSX. Returns list of raw row dicts."""
    ext = os.path.splitext(file_path)[1].lower()
    rows = []

    if ext == ".csv":
        # Try common encodings in order; many exported CSVs use UTF-8-sig or latin-1
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with open(file_path, newline="", encoding=encoding) as f:
                    reader = csv.DictReader(f)
                    rows = [dict(row) for row in reader]
                break
            except UnicodeDecodeError:
                rows = []
                continue

    elif ext in (".xlsx", ".xls"):
        if not XLSX_SUPPORTED:
            raise ImportError("openpyxl is required for XLSX support. Run: pip install openpyxl")
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({h: v for h, v in zip(headers, row)})

    else:
        raise ValueError(f"Unsupported file type: {ext}. Use .csv or .xlsx")

    return rows


def validate_jobs(rows: list[dict], logger=None) -> tuple[list[dict], list[dict]]:
    """
    Validate and clean raw rows.
    Returns (clean_jobs, error_records).
    Each clean job is a normalized dict. Each error record includes an 'errors' list.
    """
    clean = []
    errors = []

    for i, row in enumerate(rows):
        row_errors = []

        # --- Required field checks ---
        for field in REQUIRED_FIELDS:
            if field not in row or row.get(field) is None or str(row.get(field, "")).strip() == "":
                row_errors.append(f"Missing required field: '{field}'")

        if row_errors:
            # If critical ID fields missing, short-circuit
            row["errors"] = row_errors
            errors.append(row)
            if logger:
                logger.warning(f"Row {i+1} [{row.get('job_id', 'NO_ID')}]: {'; '.join(row_errors)}")
            continue

        job_id = str(row["job_id"]).strip()

        # --- Date parsing ---
        start_date = parse_date(str(row.get("start_date", "")))
        planned_end_date = parse_date(str(row.get("planned_end_date", "")))
        actual_end_str = str(row.get("actual_end_date", "")).strip()
        actual_end_date = parse_date(actual_end_str) if actual_end_str else None

        if start_date is None:
            row_errors.append(f"Invalid start_date: '{row.get('start_date')}'")
        if planned_end_date is None:
            row_errors.append(f"Invalid planned_end_date: '{row.get('planned_end_date')}'")
        if actual_end_str and actual_end_date is None:
            row_errors.append(f"Invalid actual_end_date: '{actual_end_str}'")

        if start_date and planned_end_date and planned_end_date < start_date:
            row_errors.append(
                f"planned_end_date ({planned_end_date}) is before start_date ({start_date})"
            )

        # --- Status check ---
        status = str(row.get("status", "")).strip()
        if status not in VALID_STATUSES:
            row_errors.append(f"Invalid status: '{status}'. Must be one of {VALID_STATUSES}")

        # --- Numeric checks ---
        try:
            markout_issues = int(row.get("markout_issues", 0))
            if markout_issues < 0:
                row_errors.append(f"markout_issues must be >= 0 (got {markout_issues})")
        except (ValueError, TypeError):
            row_errors.append(f"markout_issues must be an integer: '{row.get('markout_issues')}'")
            markout_issues = 0

        try:
            inspections_failed = int(row.get("inspections_failed", 0))
            if inspections_failed < 0:
                row_errors.append(f"inspections_failed must be >= 0 (got {inspections_failed})")
        except (ValueError, TypeError):
            row_errors.append(f"inspections_failed must be an integer: '{row.get('inspections_failed')}'")
            inspections_failed = 0

        # --- Bool check ---
        markout_required = parse_bool(row.get("markout_required"))
        if markout_required is None:
            row_errors.append(f"markout_required must be True/False: '{row.get('markout_required')}'")
            markout_required = False

        if row_errors:
            row["errors"] = row_errors
            errors.append(row)
            if logger:
                logger.warning(f"Row {i+1} [{job_id}]: {'; '.join(row_errors)}")
            continue

        # --- Clean record ---
        clean.append({
            "job_id": job_id,
            "utility_owner": str(row["utility_owner"]).strip(),
            "contractor": str(row["contractor"]).strip(),
            "scope_type": str(row["scope_type"]).strip(),
            "region": str(row["region"]).strip(),
            "start_date": start_date,
            "planned_end_date": planned_end_date,
            "actual_end_date": actual_end_date,
            "status": status,
            "markout_required": markout_required,
            "markout_issues": max(0, markout_issues),
            "inspections_failed": max(0, inspections_failed),
            "crew_type": str(row["crew_type"]).strip(),
        })

    return clean, errors


def save_clean_jobs(clean_jobs: list[dict], output_path: str = "outputs/clean_jobs.csv"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    if not clean_jobs:
        print("No clean jobs to save.")
        return
    fieldnames = list(clean_jobs[0].keys())
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for job in clean_jobs:
            row = dict(job)
            for k, v in row.items():
                if hasattr(v, "isoformat"):
                    row[k] = v.isoformat()
            writer.writerow(row)
    print(f"  Clean jobs saved → {output_path} ({len(clean_jobs)} records)")


def log_validation_errors(error_records: list[dict], log_path: str = "logs/validation_errors.log"):
    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    with open(log_path, "a") as f:
        f.write(f"\n--- Validation Run: {datetime.now().isoformat()} ---\n")
        if not error_records:
            f.write("No validation errors found.\n")
            return
        for rec in error_records:
            job_id = rec.get("job_id") or "UNKNOWN"
            errs = rec.get("errors", ["Unknown error"])
            f.write(f"  [{job_id}] {' | '.join(errs)}\n")
    print(f"  Validation errors logged → {log_path} ({len(error_records)} invalid records)")


def run_ingestion(file_path: str) -> list[dict]:
    """Full ingestion pipeline. Returns clean jobs."""
    logger = setup_logging()
    print(f"\n[INGESTION] Loading: {file_path}")
    rows = load_jobs(file_path)
    print(f"  Rows loaded: {len(rows)}")

    clean, errors = validate_jobs(rows, logger=logger)
    print(f"  Valid records: {len(clean)} | Invalid: {len(errors)}")

    save_clean_jobs(clean)
    log_validation_errors(errors)

    return clean
